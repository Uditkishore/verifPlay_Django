import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Add headers
headers = ['Register Name', 'Offset', 'Read/Write', 'Fields', 'Default value', 'Reset value', 'Description']
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Add data
data = [
    ['CTRL_REG', '0x00', 'RW', 'ENABLE [0:0]', 0, 0, 'Control Register'],
    ['', '', '', 'MODE [2:1]', 0, 0, 'Operating Mode'],
    ['', '', '', 'INTR_EN [3:3]', 0, 0, 'Interrupt Enable'],
    ['STATUS_REG', '0x04', 'RO', 'BUSY [0:0]', 0, 0, 'Busy Status'],
    ['', '', '', 'ERROR [1:1]', 0, 0, 'Error Status'],
    ['DATA_REG', '0x08', 'RW', 'DATA [31:0]', 0, 0, 'Data Register']
]

for row_num, row_data in enumerate(data, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook
wb.save('test_registers.xlsx')
print("Test Excel file created: test_registers.xlsx")
